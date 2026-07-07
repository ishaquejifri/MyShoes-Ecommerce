from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from .decorators import admin_required
from accounts.models import CustomUser
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, F
from django.db.models.functions import TruncMonth, TruncYear
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .decorators import admin_required
from django.utils import timezone
from django.http import HttpResponse
from orders.models import Order, OrderItem
from products.models import Product
from category.models import Category
import json
from datetime import timedelta, date 
import io

User = get_user_model()


@never_cache
def admin_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('admin_dashboard')

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user:
            if user.is_staff:
                login(request, user)
                return redirect('admin_dashboard')
            else:
                messages.error(request, 'You are not autherized to access admin panel')
        else:
            messages.error(request, 'Invalid username or password')

    return render(request, "admin_login.html")


def admin_logout(request):
    storage = messages.get_messages(request)
    for _ in storage:
        pass
    logout(request)
    request.session.flush()
    return redirect('admin_login')


@admin_required
@never_cache
@login_required
@staff_member_required(login_url='admin_login')
def admin_dashboard(request):
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)

    # --- Summary Cards ---
    total_revenue = Order.objects.filter(
        status__in=['delivered', 'shipped', 'out_for_delivery', 'processing', 'confirmed']
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    daily_revenue = Order.objects.filter(
        created_at__date=today,
        status__in=['delivered', 'shipped', 'out_for_delivery', 'processing', 'confirmed']
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    active_users = User.objects.filter(is_active=True, is_staff=False).count()
    pending_orders = Order.objects.filter(status='pending').count()
    total_orders = Order.objects.count()

    # --- Monthly Revenue for Chart (last 12 months) ---
    twelve_months_ago = today - timedelta(days=365)
    monthly_data = (
        Order.objects
        .filter(created_at__date__gte=twelve_months_ago,
                status__in=['delivered', 'shipped', 'out_for_delivery', 'processing', 'confirmed'])
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(revenue=Sum('total_amount'))
        .order_by('month')
    )
    chart_labels = [entry['month'].strftime('%b %Y') for entry in monthly_data]
    chart_data = [float(entry['revenue']) for entry in monthly_data]

    # --- Top 10 Best Selling Products ---
    top_products = (
        OrderItem.objects
        .values('product__id', 'product__product_name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:10]
    )

    # --- Top 10 Best Selling Categories ---
    top_categories = (
        OrderItem.objects
        .values('product__category__id', 'product__category__name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:10]
    )

    # --- Top 10 Best Selling Brands ---
    top_brands = (
        OrderItem.objects
        .values('product__brand')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:10]
    )

    # --- Recent Orders ---
    recent_orders = Order.objects.select_related('user').order_by('-created_at')[:10]

    context = {
        'total_revenue': total_revenue,
        'daily_revenue': daily_revenue,
        'active_users': active_users,
        'pending_orders': pending_orders,
        'total_orders': total_orders,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'top_products': top_products,
        'top_categories': top_categories,
        'top_brands': top_brands,
        'recent_orders': recent_orders,
    }
    return render(request, "dashboard.html", context)


@never_cache
@admin_required
@login_required(login_url='admin_login')
def user_list(request):
    search_query = request.GET.get('search', '')
    users = User.objects.filter(is_superuser=False).annotate(total_orders=Count('orders')).order_by('-date_joined')

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    paginator = Paginator(users, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query
    }
    return render(request, "user_list.html", context)


@admin_required
@login_required(login_url='admin_login')
def admin_sales_report(request):
    filter_type = request.GET.get('filter', 'monthly')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    today = timezone.now().date()

    if filter_type == 'daily':
        start_date = today
        end_date = today
    elif filter_type == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif filter_type == 'custom' and start_date_str and end_date_str:
        from datetime import datetime
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:  # monthly default
        start_date = today.replace(day=1)
        end_date = today

    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).exclude(status='cancelled').order_by('-created_at')

    total_sales_count = orders.count()
    total_order_amount = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    total_discount = orders.aggregate(total=Sum('discount_amount'))['total'] or 0
    total_coupon_discount = orders.aggregate(total=Sum('coupon_discount'))['total'] or 0
    overall_discount = (total_discount or 0) + (total_coupon_discount or 0)

    paginator = Paginator(orders, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'orders': page_obj,
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'start_date_str': start_date_str,
        'end_date_str': end_date_str,
        'total_sales_count': total_sales_count,
        'total_order_amount': total_order_amount,
        'total_discount': total_discount,
        'total_coupon_discount': total_coupon_discount,
        'overall_discount': overall_discount,
        'page_obj': page_obj,
        'filter_choices': [
            ('daily', 'Today'),
            ('weekly', 'This Week'),
            ('monthly', 'This Month'),
            ('yearly', 'This Year'),
        ],
    }
    return render(request, 'sales_report.html', context)


@admin_required
@login_required(login_url='admin_login')
def download_sales_report_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    filter_type = request.GET.get('filter', 'monthly')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    today = timezone.now().date()

    if filter_type == 'daily':
        start_date = today
        end_date = today
    elif filter_type == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif filter_type == 'custom' and start_date_str and end_date_str:
        from datetime import datetime
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = today.replace(day=1)
        end_date = today

    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).exclude(status='cancelled').order_by('-created_at')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 alignment=TA_CENTER, fontSize=18,
                                 textColor=colors.HexColor('#10231b'))
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'],
                                    alignment=TA_CENTER, fontSize=11,
                                    textColor=colors.grey)

    elements.append(Paragraph("Lotto Shoes - Sales Report", title_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 16))

    # Summary row
    total_amount = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    total_discount = orders.aggregate(total=Sum('discount_amount'))['total'] or 0
    total_coupon = orders.aggregate(total=Sum('coupon_discount'))['total'] or 0
    summary_data = [
        ['Total Orders', 'Total Amount', 'Offer Discount', 'Coupon Discount', 'Net Revenue'],
        [
            str(orders.count()),
            f'Rs. {total_amount:,.2f}',
            f'Rs. {total_discount:,.2f}',
            f'Rs. {total_coupon:,.2f}',
            f'Rs. {float(total_amount):,.2f}',
        ]
    ]
    summary_table = Table(summary_data, colWidths=[1.8*inch]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10231b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e8f5ef')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5faf7')]),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Orders table
    header = ['Order ID', 'Date', 'Customer', 'Payment', 'Status', 'Subtotal', 'Discount', 'Coupon', 'Total']
    data = [header]
    for order in orders:
        data.append([
            order.order_id,
            order.created_at.strftime('%d-%m-%Y'),
            order.full_name[:20],
            order.payment_method.upper(),
            order.status.replace('_', ' ').title(),
            f'Rs.{order.sub_total:,.0f}',
            f'Rs.{order.discount_amount:,.0f}',
            f'Rs.{order.coupon_discount:,.0f}',
            f'Rs.{order.total_amount:,.0f}',
        ])

    col_widths = [0.9*inch, 0.9*inch, 1.5*inch, 0.8*inch, 1.1*inch, 0.9*inch, 0.8*inch, 0.8*inch, 0.9*inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#224939')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f8f4')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.pdf"'
    return response


@admin_required
@login_required(login_url='admin_login')
def download_sales_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filter_type = request.GET.get('filter', 'monthly')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    today = timezone.now().date()

    if filter_type == 'daily':
        start_date = today
        end_date = today
    elif filter_type == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
    elif filter_type == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif filter_type == 'custom' and start_date_str and end_date_str:
        from datetime import datetime
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = today.replace(day=1)
        end_date = today

    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).exclude(status='cancelled').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    header_fill = PatternFill(start_color='10231b', end_color='10231b', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    summary_fill = PatternFill(start_color='224939', end_color='224939', fill_type='solid')
    alt_fill = PatternFill(start_color='f0f8f4', end_color='f0f8f4', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )

    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Lotto Shoes - Sales Report'
    ws['A1'].font = Font(bold=True, size=16, color='10231b')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:I2')
    ws['A2'] = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    # Summary
    total_amount = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    total_discount = orders.aggregate(total=Sum('discount_amount'))['total'] or 0
    total_coupon = orders.aggregate(total=Sum('coupon_discount'))['total'] or 0

    ws.append([])
    summary_headers = ['Total Orders', 'Total Amount', 'Offer Discount', 'Coupon Discount', 'Net Revenue']
    ws.append(summary_headers)
    for col_idx, val in enumerate(summary_headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = summary_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = center

    summary_values = [
        orders.count(),
        float(total_amount),
        float(total_discount),
        float(total_coupon),
        float(total_amount)
    ]
    ws.append(summary_values)
    for col_idx in range(1, 6):
        ws.cell(row=ws.max_row, column=col_idx).alignment = center

    ws.append([])

    # Order headers
    headers = ['Order ID', 'Date', 'Customer', 'Email', 'Payment', 'Status', 'Subtotal', 'Offer Disc.', 'Coupon Disc.', 'Total']
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    for idx, order in enumerate(orders):
        row = [
            order.order_id,
            order.created_at.strftime('%d-%m-%Y'),
            order.full_name,
            order.user.email,
            order.payment_method.upper(),
            order.status.replace('_', ' ').title(),
            float(order.sub_total),
            float(order.discount_amount),
            float(order.coupon_discount),
            float(order.total_amount),
        ]
        ws.append(row)
        fill = alt_fill if idx % 2 == 1 else None
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            if fill:
                cell.fill = fill
            cell.border = thin_border
            cell.alignment = center

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.xlsx"'
    return response


@login_required(login_url='admin_login')
def toggle_block_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if user.is_block:
        user.is_block = False
        messages.success(request, 'User unblocked successfully')
    else:
        user.is_block = True
        messages.success(request, 'User blocked successfully')

    user.save()
    return redirect('user_list')
